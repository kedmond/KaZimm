# -*- coding: utf-8 -*-
"""
Created on Thu Apr 10 10:07:26 2014

@author: kedmond
"""

import glob
import numpy as np
import counter as cnt
import os

import openpyxl as xl
import csv
import time
import re

# 10.April.2014:
# Instead of Excel, let's just put all the data in single CSV files with 
# multiple columns. Then sort the experiments by folder.
# We need to move the data writing functions to a higher level so that we can 
# manage things more easily.

freqs = np.array([50, 100, 200, 400, 600, 800, 1000, 2000])
dur = np.array([ 600, 420, 300, 240, 120,  60,   30,   10])
sr = np.array([  128, 128, 256, 512, 512, 1024, 1024, 2048])

numMeas = 5

i = 0

# make an empty array to save the data:
dataStack = np.zeros((sr[i]*dur[i], numMeas))

# normally data is acquired like this:
for run in range(numMeas):
    dataStack[:, i] = cnt.retrieveData(sr[0], dur[0])

# *** Create directory structure for data from experiments ***

rootDir = '/home/kedmond/Desktop/'
dateDir = time.strftime('%d%b%Y') + '/'
expStr = 'Experiment'
expNum = 1

# Look for other "Experiment" directories, and create the next one 

if not os.path.exists(rootDir + dateDir + expStr + str(expNum)):
    os.makedirs(rootDir + dateDir + expStr + str(expNum))
    # going forward, this is where to store data
    directory = rootDir + dateDir + expStr + str(expNum)
else:
    # get a list of directories that start with 'Experiment', count them, and enumerate
    fns = glob.glob(rootDir + dateDir + expStr + '*')
    fns.sort()
    
    expNum = int(re.findall(r'\d+', fns[-1][-11:])[0]) + 1
    
    os.makedirs(rootDir + dateDir + expStr + str(expNum))
    
    # going forward, this is where to store data
    directory = rootDir + dateDir + expStr + str(expNum)

# *** *** *** *** *** *** *** *** *** ****** *** *** *** *** *** 

fname = 'rotorSpeeds-'+ str(freqs[i]).zfill(4) +'.txt'

np.savetxt(directory + '/' + fname, dataStack, delimiter=', ', 
           header='duration: ' + str(dur[0]) + ', sampRate: ' + str(sr[0]))







# Excel stuff

dir = '/media/HDD/Data/ZimmViscometer/Data/'

fns = glob.glob(dir+'BTF2/31.Mar.2014/'+'14-03*.txt')
data = np.genfromtxt(fns[0], delimiter=',', skiprows=2)
data[:, 1] *= 180/np.pi

# pure D2O/H2O mixture, 1.052 g/mL:
fn2 = '/media/HDD/Data/ZimmViscometer/Data/Calibrations/13-10-01-16-37-56_calibration-H2O+D2O-1.052-2.txt'
cal2 = np.genfromtxt(fn2, skiprows=2, delimiter=',')


# let's make an Excel workbook and save data to it

fns = glob.glob(dir+'BTF2/9.April.2014/Set 1/'+'data*.txt')
fns.sort()

wb = xl.Workbook(optimized_write=True)

ws1 = wb.create_sheet(title='averages')

#for irow in xrange(data[:, 0].size):
#    ws1.append([data[irow, 1], data[irow, 2]])

for i in xrange(np.size(fns)):    
    vals = np.genfromtxt(fns[i], delimiter=',', skip_header=1)[:, 1]
    r = csv.reader(open(fns[i]))
    line1 = r.next()
    line1 = line1[0] + line1[1]
    hdrs = [int(s) for s in line1.split() if s.isdigit()]
    
    wsd = wb.create_sheet(title=str(hdrs[1])+', '+str(hdrs[0]))
    
    for irow in xrange( int(vals.size/100) ):
        wsd.append( [vals[irow]] )

wb.save('file.xlsx')


for i in xrange(1):    
    vals = np.genfromtxt(fns[i], delimiter=',', skip_header=1)[:, 1]
    r = csv.reader(open(fns[i]))
    line1 = r.next()
    line1 = line1[0] + line1[1]
    hdrs = [int(s) for s in line1.split() if s.isdigit()]
    
    wsd = wb.create_sheet(title=str(hdrs[1])+', '+str(hdrs[0]))
    
    for irow in xrange( int(vals.size) ):
        wsd.append( [vals[irow]] )

wb.save('file.xlsx')



from openpyxl.cell import get_column_letter

wb = xl.workbook()
dest_filename = r'empty_book.xlsx'
ws = wb.active
ws.title = "range names"

for col_idx in xrange(1, 40):
    col = get_column_letter(col_idx)
    for row in xrange(1, 600):
        ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)

ws = wb.create_sheet()
ws.title = 'Pi'
ws['F5'] = 3.14
wb.save(filename=dest_filename)


